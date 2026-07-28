from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def criar_planilha_base():
    wb = Workbook()

    aba = wb.active

    while True:
        nome_arquivo = input('Informe o nome da Planilha (a extensão deve terminar com ".xlsx"): ')
        nome_aba = input('Informe o nome da aba: ').strip()

        if not nome_arquivo or not nome_aba:
            print('Preencha o campo corretamente.')
            continue     
        
        if not nome_arquivo.endswith('.xlsx'):
            nome_arquivo += '.xlsx'
        
        break

    aba.title = nome_aba
    nomes_colunas = []

    while True:
        try:
            qtd_colunas = int(input('Informe a quantidade de colunas: ').strip())

            if qtd_colunas <= 0:
                print('Preencha o campo com um número inteiro válido: ')
                continue 
            break      
        except ValueError:
            print('Preencha o campo com um número inteiro válido: ')
            continue
    
    for indice, coluna in enumerate(range(1, qtd_colunas + 1), start=1):
        nomes = input(f'Informe o nome da coluna {indice}: ')
        nomes_colunas.append(nomes)

    aba.append(nomes_colunas)

    while True:
        try:
            linhas = int(input('Informe a qtd de linhas da planilha: ').strip())

            if linhas <= 0:
                print('Preencha o campo com um número inteiro válido: ')
                continue 
            break      
        except ValueError:
            print('Preencha o campo com um número inteiro válido: ')
            continue

    for indice, linha in enumerate(range(1, linhas + 1), start=1):
        print(f'\nPreenchendo a linha {indice}: ')

        dados_linha = []

        for coluna in nomes_colunas:
            valor = input(f'Informe o valor de {coluna}: ')
            dados_linha.append(valor)
    
        aba.append(dados_linha)
    
    wb.save(nome_arquivo)

    print(f'Planilha {nome_arquivo} Criada com Sucesso!')
        
    return nome_arquivo, nome_aba

def gerar_relatorio(nome_arquivo, nome_aba):
    wb = load_workbook(nome_arquivo)

    aba_vendas = wb[nome_aba]

    if 'Relatório' in wb.sheetnames:
        aba_antiga = wb['Relatório']
        wb.remove(aba_antiga)
    
    relatorio = wb.create_sheet('Relatório')

    for linhas in aba_vendas.iter_rows(values_only=True):
        relatorio.append(list(linhas))
    
    wb.save(nome_arquivo)

    print('Aba Relatório Criada com sucesso!')

def formatar_relatorio(nome_arquivo):
    wb = load_workbook(nome_arquivo)

    relatorio = wb['Relatório']

    for celula in relatorio[1]:
        celula.font = Font(bold=True)
    
    for numero_coluna in range(1, relatorio.max_column + 1):
        letra_coluna = get_column_letter(numero_coluna)
        relatorio.column_dimensions[letra_coluna].width = 20
    
    relatorio.freeze_panes = 'A2' 

    wb.save(nome_arquivo)

    print('Aba Relatório formatada com sucesso!')

def criar_resumo(nome_arquivo):
    wb = load_workbook(nome_arquivo)

    relatorio = wb['Relatório']

    columns = []

    for celula in relatorio[1]:
        columns.append(celula.value)

    for indice, coluna in enumerate(columns, start=1):
        print(f'{indice} - {coluna}')
    
    while True:
        try: 
            escolha = int(input('Escolha uma coluna para gerar o resumo: ').strip())

            if escolha <= 0 or escolha > len(columns):
                print('Preencha o Campo Corretamente!')
                continue 

            break
        except ValueError:
            print('Preencha o Campo Corretamente!')
            continue 
    
    letra_coluna = get_column_letter(escolha)

    resumo = {}

    for linha in range(2, relatorio.max_row + 1):
        valor = relatorio[f'{letra_coluna}{linha}'].value

        if valor not in resumo:
            resumo[valor] = 0
        
        resumo[valor] += 1
    
    if 'Resumo' in wb.sheetnames:
        aba_antiga = wb['Resumo']
        wb.remove(aba_antiga)
    
    aba_resumo = wb.create_sheet('Resumo')

    aba_resumo.append(['Valor', 'Quantidade'])

    for valor, quantidade in resumo.items():
        aba_resumo.append([valor, quantidade])
    
    for celula in aba_resumo[1]:
        celula.font = Font(bold=True)
    
    aba_resumo.column_dimensions['A'].width = 25
    aba_resumo.column_dimensions['B'].width = 15

    wb.save(nome_arquivo)

    print('Aba Resumo criada com sucesso!')

def main():
    nome_arquivo, nome_aba = criar_planilha_base()
    gerar_relatorio(nome_arquivo, nome_aba)
    formatar_relatorio(nome_arquivo)
    criar_resumo(nome_arquivo)


if __name__ == '__main__':
    main()