from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

def carregar_cadastros(caminho_arquivo):

    caminho = Path(caminho_arquivo)
    
    if not caminho.exists():
        return []

    wb = load_workbook(caminho)

    aba = wb["Cadastros"]

    cadastros = []


    for linha in aba.iter_rows(min_row=2, values_only=True):

        nome, email, telefone, empresa, cargo, cidade, estado, observacoes = linha

        cadastro = {
            "nome": nome, 
            "email": email, 
            "telefone": telefone, 
            "empresa": empresa,
            "cargo": cargo,
            "cidade": cidade, 
            "estado": estado,
            "observacoes": observacoes,
        }

        cadastros.append(cadastro)

    wb.close()

    return cadastros

def validar_cadastro(cadastro):

    campos_obrigatorios = [
        "nome",
        "email",
        "telefone",
        "empresa",
        "cargo",
        "cidade",
        "estado"
    ]

    campos_invalidos = []

    for campo in campos_obrigatorios:

        valor = cadastro.get(campo, "")

        if not valor:
            campos_invalidos.append(campo)


    if campos_invalidos:
        mensagem = f"Campos não preenchidos: {', '.join(campos_invalidos)}"
        return False, mensagem

    return True, ""

def salvar_resultados(resultados, invalidos, caminho_saida):

    wb = Workbook()

    aba_resultados = wb.active

    aba_resultados.title = "Resultados"
    aba_invalidos = wb.create_sheet("Inválidos")

    aba_resultados.append(["nome", "status", "mensagem"])

    for linha in aba_resultados[1]:
        linha.font = Font(bold=True)

    aba_invalidos.append([
            "nome",
            "email",
            "telefone",
            "empresa",
            "cargo",
            "cidade",
            "estado", 
            "erro"]
    )

    for linha in aba_invalidos[1]:
            linha.font = Font(bold=True)


    for resultado in resultados:
        aba_resultados.append([resultado["nome"], resultado["status"], resultado["mensagem"]])

    for cadastro in invalidos:
        aba_invalidos.append([
            cadastro.get("nome", ""),
            cadastro.get("email", ""),
            cadastro.get("telefone", ""),
            cadastro.get("empresa", ""),
            cadastro.get("cargo", ""),
            cadastro.get("cidade", ""),
            cadastro.get("estado", ""), 
            cadastro.get("erro", "")
        ])

    wb.save(caminho_saida)

    