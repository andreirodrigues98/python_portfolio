from openpyxl import load_workbook

def carregar_planilha():

    while True: 
        nome_arquivo = input('Informe o nome do arquivo(.xlsx): ').strip()
        nome_aba = input('Informe o nome da aba do arquivo: ').strip()

        if not nome_aba or not nome_arquivo:
            print('O campo deve ser preenchido corretamente!')
            continue 

        if not nome_arquivo.endswith('.xlsx'):
            nome_arquivo += '.xlsx'
            
        break

    wb = load_workbook(nome_arquivo)

    if nome_aba in wb.sheetnames:
        aba = wb[nome_aba]
        return wb, aba, nome_arquivo
    else:
        print(f'Não existe uma aba com este nome({nome_aba}) na planilha!')
        
def listar_colunas(aba):

    print('\n========== COLUNAS DA PLANILHA ==========')

    for indice, celula in enumerate(aba[1], start=1):
        print(f'{indice} - {celula.value}')

def escolher_coluna(aba):

    while True:

        try:
            escolha = int(input('Escolha uma das colunas acima: ').strip())

            if escolha <= 0 or escolha > len(aba[1]):
                print('Selecione um numero válido')
                continue 

            break

        except ValueError:
            print('Selecione um numero válido')
            continue
    
    return escolha

def analise_coluna_escolhida(aba, coluna_escolhida):
    lista_numerica = []

    for linha in range(2, aba.max_row + 1):
        valor = aba.cell(row=linha, column=coluna_escolhida).value

        try:  
            numero = float(valor)
            
            lista_numerica.append(numero)

        except (ValueError, TypeError):
           continue


    if not lista_numerica:
        print('A lista esta vazia, não há nenhum numero para analisar!')
        return 
    
    quantidade =  len(lista_numerica)
    soma = sum(lista_numerica)
    maior = max(lista_numerica)
    menor = min(lista_numerica)
    media = soma / quantidade

    dados = {
        "quantidade": quantidade,
        "soma": soma,
        "media": media,
        "maior": maior,
        "menor": menor
    }

    return dados

def salvar_relatorio(dados, nome_arquivo):

    if nome_arquivo.endswith('.xlsx'):
        nome_arquivo = nome_arquivo.replace(".xlsx", "")
    
    with open(f'{nome_arquivo}.txt', 'w', encoding='utf-8') as arquivo:
        arquivo.write("========== Relatório ==========\n")
        arquivo.write(f"Planilha analisada: {nome_arquivo} (.xlsx)\n\n")

        arquivo.write(f"Quantidade de numeros encontrados: {dados['quantidade']:.2f}\n")
        arquivo.write(f"Soma dos numeros encontrados: {dados['soma']:.2f}\n")
        arquivo.write(f"Média dos numeros encontrados: {dados['media']:.2f}\n")
        arquivo.write(f"Maior numero encontrado: {dados['maior']:.2f}\n")
        arquivo.write(f"Menor numero encontrado: {dados['menor']:.2f}")

    
    print("Relatorio gerado com sucesso!")

def main():
    resultado = carregar_planilha()

    if not resultado:
        return 
    
    wb, aba, nome_arquivo = resultado
    listar_colunas(aba)
    coluna_escolhida = escolher_coluna(aba)
    dados = analise_coluna_escolhida(aba, coluna_escolhida)

    if not dados:
        return 
    
    salvar_relatorio(dados, nome_arquivo)

if __name__ == '__main__':
    main()